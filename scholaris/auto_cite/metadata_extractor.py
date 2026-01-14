#!/usr/bin/env python3
"""
Hybrid Metadata Extractor for PDFs.

Combines multiple extraction methods:
1. pdf2bib - DOI/ISBN lookup from PDF metadata
2. Gemini Vision - OCR of title pages for visual extraction
3. PDF internal metadata
4. Filename parsing - Extract hints from filename patterns

Intelligently merges results to produce the most accurate metadata.
"""

import base64
import csv
import json
import logging
import re
import time
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import List, Optional, Dict, Any

import fitz  # PyMuPDF

logger = logging.getLogger(__name__)


@dataclass
class ExtractedMetadata:
    """Extracted metadata from a PDF."""
    title: str = ""
    authors: List[str] = field(default_factory=list)
    year: Optional[int] = None
    source: str = ""  # Journal, publisher, etc.
    doi: str = ""
    isbn: str = ""
    abstract: str = ""
    citation_key: str = ""
    volume: str = ""
    issue: str = ""
    pages: str = ""
    entry_type: str = "article"
    confidence: float = 0.0  # 0-1 confidence in extraction
    extraction_method: str = ""  # Which method(s) produced this

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return asdict(self)

    def to_bibtex(self) -> str:
        """Generate BibTeX entry."""
        key = self.citation_key or self._generate_key()

        # Determine entry type
        entry_type = self.entry_type
        if self.isbn and entry_type == "article":
            entry_type = "book"
        elif "thesis" in self.source.lower():
            entry_type = "mastersthesis"

        lines = [f"@{entry_type}{{{key},"]

        if self.title:
            lines.append(f"  title = {{{self.title}}},")
        if self.authors:
            lines.append(f"  author = {{{' and '.join(self.authors)}}},")
        if self.year:
            lines.append(f"  year = {{{self.year}}},")
        if self.source:
            if entry_type == "article":
                lines.append(f"  journal = {{{self.source}}},")
            else:
                lines.append(f"  publisher = {{{self.source}}},")
        if self.volume:
            lines.append(f"  volume = {{{self.volume}}},")
        if self.issue:
            lines.append(f"  number = {{{self.issue}}},")
        if self.pages:
            lines.append(f"  pages = {{{self.pages}}},")
        if self.doi:
            lines.append(f"  doi = {{{self.doi}}},")
        if self.isbn:
            lines.append(f"  isbn = {{{self.isbn}}},")

        lines.append("}")
        return "\n".join(lines)

    def _generate_key(self) -> str:
        """Generate a citation key."""
        if self.authors and self.year:
            first_author = self.authors[0].split()[-1].lower()
            first_author = re.sub(r'[^a-z]', '', first_author)
            return f"{first_author}{self.year}"
        elif self.title and self.year:
            first_word = re.sub(r'[^a-z]', '', self.title.split()[0].lower())
            return f"{first_word}{self.year}"
        return "unknown"


class HybridMetadataExtractor:
    """
    Hybrid metadata extractor combining multiple methods.

    Priority order:
    1. pdf2bib (DOI/ISBN lookup) - highest confidence when found
    2. Gemini Vision (title page OCR) - good for scanned/academic PDFs
    3. PDF internal metadata - often incomplete/wrong
    4. Filename parsing - fallback hints

    The results are intelligently merged with higher confidence sources
    taking precedence.
    """

    def __init__(self, gemini_api_key: Optional[str] = None, model_name: str = "gemini-2.0-flash"):
        """Initialize extractor with optional Gemini API key."""
        self.gemini_api_key = gemini_api_key
        self.model_name = model_name
        self._gemini_model = None

    def extract(
        self,
        pdf_path: str,
        use_vision: bool = True,
        existing_metadata: Optional[Dict[str, Any]] = None
    ) -> ExtractedMetadata:
        """
        Extract metadata from PDF using hybrid approach.

        Args:
            pdf_path: Path to PDF file
            use_vision: Whether to use Gemini Vision (requires API key)
            existing_metadata: Pre-existing metadata to incorporate

        Returns:
            ExtractedMetadata with best available data
        """
        pdf_path = Path(pdf_path)

        results = []
        methods_used = []

        # Add existing metadata if provided
        if existing_metadata:
            existing = ExtractedMetadata(
                title=existing_metadata.get("title", ""),
                authors=existing_metadata.get("authors", []),
                year=existing_metadata.get("year"),
                source=existing_metadata.get("source", ""),
                citation_key=existing_metadata.get("citation_key", ""),
                doi=existing_metadata.get("doi", ""),
                confidence=0.95,  # User-provided data is highest confidence
                extraction_method="user_provided"
            )
            if existing.title or existing.authors:
                results.append(existing)
                methods_used.append("user_provided")

        # Method 1: Try pdf2bib first (most reliable for DOI/ISBN)
        pdf2bib_meta = self._extract_pdf2bib(pdf_path)
        if pdf2bib_meta and pdf2bib_meta.confidence > 0.5:
            results.append(pdf2bib_meta)
            methods_used.append("pdf2bib")
            logger.debug(f"pdf2bib extracted: {pdf2bib_meta.title[:50] if pdf2bib_meta.title else 'N/A'}...")

        # Method 2: Gemini Vision OCR
        if use_vision and self.gemini_api_key:
            vision_meta = self._extract_gemini_vision(pdf_path)
            if vision_meta and vision_meta.confidence > 0.3:
                results.append(vision_meta)
                methods_used.append("gemini_vision")
                logger.debug(f"Gemini Vision extracted: {vision_meta.title[:50] if vision_meta.title else 'N/A'}...")

        # Method 3: PDF internal metadata
        internal_meta = self._extract_pdf_metadata(pdf_path)
        if internal_meta and internal_meta.confidence > 0.2:
            results.append(internal_meta)
            methods_used.append("pdf_metadata")

        # Method 4: Filename parsing (fallback)
        filename_meta = self._extract_from_filename(pdf_path)
        if filename_meta:
            results.append(filename_meta)
            methods_used.append("filename")

        # Merge results intelligently
        final = self._merge_results(results)
        final.extraction_method = "+".join(methods_used) if methods_used else "none"

        # Generate citation key if not set
        if not final.citation_key:
            final.citation_key = final._generate_key()

        return final

    def _extract_pdf2bib(self, pdf_path: Path) -> Optional[ExtractedMetadata]:
        """Extract metadata using pdf2bib library."""
        try:
            from pdf2bib import pdf2bib
            result = pdf2bib.pdf2bib(str(pdf_path))

            if result and result.get("metadata"):
                meta = result["metadata"]
                authors = []
                if meta.get("author"):
                    authors = [a.strip() for a in meta["author"].split(" and ")]

                return ExtractedMetadata(
                    title=meta.get("title", ""),
                    authors=authors,
                    year=int(meta.get("year", 0)) if meta.get("year") else None,
                    source=meta.get("journal", meta.get("publisher", "")),
                    doi=meta.get("doi", ""),
                    isbn=meta.get("isbn", ""),
                    volume=meta.get("volume", ""),
                    pages=meta.get("pages", ""),
                    entry_type=meta.get("ENTRYTYPE", "article"),
                    confidence=0.9 if meta.get("doi") or meta.get("isbn") else 0.6,
                    extraction_method="pdf2bib"
                )
        except Exception as e:
            logger.debug(f"pdf2bib extraction failed: {e}")
        return None

    def _extract_gemini_vision(self, pdf_path: Path) -> Optional[ExtractedMetadata]:
        """Extract metadata using Gemini Vision OCR on title pages."""
        if not self.gemini_api_key:
            return None

        try:
            import google.generativeai as genai

            if not self._gemini_model:
                genai.configure(api_key=self.gemini_api_key)
                self._gemini_model = genai.GenerativeModel(self.model_name)

            # Render first 2-3 pages
            images = self._render_pages(pdf_path, max_pages=3)
            if not images:
                return None

            prompt = """Analyze these pages from an academic PDF and extract bibliographic metadata.
Return ONLY a valid JSON object with these fields:

{
    "title": "The full, exact title as printed",
    "authors": ["Author 1 Full Name", "Author 2 Full Name"],
    "year": 1234,
    "source": "Journal/Publisher name, volume, pages if visible",
    "doi": "10.xxxx/xxxxx if visible",
    "volume": "volume number if visible",
    "pages": "page range if visible",
    "entry_type": "article or book or thesis"
}

Rules:
- Extract the REAL title exactly as printed on the document
- List ALL authors with their full names in order
- Use the original publication year (not reprint/edition year)
- For journals include volume/issue/pages if visible
- If a field is not clearly visible, use null
- Return ONLY valid JSON, no other text or markdown"""

            content = [prompt]
            for img_bytes in images:
                content.append({
                    "mime_type": "image/png",
                    "data": img_bytes
                })

            response = self._gemini_model.generate_content(content)
            text = response.text.strip()

            # Parse JSON
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]

            data = json.loads(text)

            return ExtractedMetadata(
                title=data.get("title") or "",
                authors=data.get("authors") or [],
                year=int(data["year"]) if data.get("year") else None,
                source=data.get("source") or "",
                doi=data.get("doi") or "",
                volume=data.get("volume") or "",
                pages=data.get("pages") or "",
                entry_type=data.get("entry_type") or "article",
                confidence=0.85,
                extraction_method="gemini_vision"
            )

        except Exception as e:
            logger.debug(f"Gemini Vision extraction failed: {e}")
        return None

    def _extract_pdf_metadata(self, pdf_path: Path) -> Optional[ExtractedMetadata]:
        """Extract metadata from PDF internal metadata."""
        try:
            doc = fitz.open(str(pdf_path))
            meta = doc.metadata
            doc.close()

            if not meta:
                return None

            title = meta.get("title", "")
            author = meta.get("author", "")

            # Parse year from various date fields
            year = None
            for date_field in ["creationDate", "modDate"]:
                date_str = meta.get(date_field, "")
                if date_str:
                    match = re.search(r'D:(\d{4})', date_str)
                    if match:
                        year = int(match.group(1))
                        break

            if not title and not author:
                return None

            return ExtractedMetadata(
                title=title,
                authors=[a.strip() for a in author.split(",")] if author else [],
                year=year,
                confidence=0.4,  # PDF metadata often unreliable
                extraction_method="pdf_metadata"
            )

        except Exception as e:
            logger.debug(f"PDF metadata extraction failed: {e}")
        return None

    def _extract_from_filename(self, pdf_path: Path) -> Optional[ExtractedMetadata]:
        """Extract hints from filename patterns."""
        filename = pdf_path.stem

        # Common patterns
        patterns = [
            # Author_Title_Year.pdf
            r'^([A-Z][a-z]+)_(.+?)_(\d{4})$',
            # Author - Title (Year).pdf
            r'^([A-Z][a-z]+)\s*-\s*(.+?)\s*\((\d{4})\)$',
            # Just year in filename
            r'(\d{4})',
        ]

        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                groups = match.groups()
                if len(groups) >= 3:
                    return ExtractedMetadata(
                        title=groups[1].replace("_", " ").replace("-", " ").strip(),
                        authors=[groups[0]] if groups[0] else [],
                        year=int(groups[2]) if groups[2].isdigit() else None,
                        confidence=0.3,
                        extraction_method="filename"
                    )
                elif len(groups) == 1 and groups[0].isdigit():
                    return ExtractedMetadata(
                        year=int(groups[0]),
                        confidence=0.2,
                        extraction_method="filename"
                    )

        return None

    def _render_pages(self, pdf_path: Path, max_pages: int = 3, dpi: int = 150) -> List[bytes]:
        """Render PDF pages to PNG images."""
        images = []
        try:
            doc = fitz.open(str(pdf_path))
            num_pages = min(max_pages, len(doc))

            for i in range(num_pages):
                page = doc[i]
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)
                images.append(pix.tobytes("png"))

            doc.close()
        except Exception as e:
            logger.debug(f"Page rendering failed: {e}")

        return images

    def _merge_results(self, results: List[ExtractedMetadata]) -> ExtractedMetadata:
        """
        Intelligently merge multiple extraction results.

        Higher confidence sources take precedence for each field.
        """
        if not results:
            return ExtractedMetadata()

        if len(results) == 1:
            return results[0]

        # Sort by confidence (highest first)
        results = sorted(results, key=lambda x: x.confidence, reverse=True)

        merged = ExtractedMetadata()

        # For each field, take from highest confidence source that has it
        for r in results:
            if r.title and not merged.title:
                merged.title = r.title
            if r.authors and len(r.authors) > len(merged.authors):
                merged.authors = r.authors
            if r.year and not merged.year:
                merged.year = r.year
            if r.source and not merged.source:
                merged.source = r.source
            if r.doi and not merged.doi:
                merged.doi = r.doi
            if r.isbn and not merged.isbn:
                merged.isbn = r.isbn
            if r.volume and not merged.volume:
                merged.volume = r.volume
            if r.pages and not merged.pages:
                merged.pages = r.pages
            if r.entry_type != "article" and merged.entry_type == "article":
                merged.entry_type = r.entry_type
            if r.abstract and len(r.abstract) > len(merged.abstract):
                merged.abstract = r.abstract
            if r.citation_key and not merged.citation_key:
                merged.citation_key = r.citation_key

        # Average confidence
        merged.confidence = sum(r.confidence for r in results) / len(results)

        return merged


class BibliographyExporter:
    """Export bibliography metadata to various formats."""

    @staticmethod
    def to_bibtex(entries: List[ExtractedMetadata], output_path: str) -> str:
        """
        Export to BibTeX format.

        Args:
            entries: List of ExtractedMetadata objects
            output_path: Output file path (.bib)

        Returns:
            BibTeX content as string
        """
        bibtex_entries = []
        for entry in sorted(entries, key=lambda x: x.year or 9999):
            bibtex_entries.append(entry.to_bibtex())

        content = "\n\n".join(bibtex_entries)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)

        return content

    @staticmethod
    def to_csv(entries: List[ExtractedMetadata], output_path: str) -> None:
        """
        Export to CSV format.

        Args:
            entries: List of ExtractedMetadata objects
            output_path: Output file path (.csv)
        """
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "citation_key", "year", "title", "authors", "source",
                "doi", "isbn", "volume", "pages", "entry_type"
            ])

            for entry in sorted(entries, key=lambda x: x.year or 9999):
                writer.writerow([
                    entry.citation_key,
                    entry.year or "",
                    entry.title,
                    "; ".join(entry.authors),
                    entry.source,
                    entry.doi,
                    entry.isbn,
                    entry.volume,
                    entry.pages,
                    entry.entry_type
                ])

    @staticmethod
    def to_xlsx(entries: List[ExtractedMetadata], output_path: str) -> None:
        """
        Export to Excel format.

        Args:
            entries: List of ExtractedMetadata objects
            output_path: Output file path (.xlsx)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl required for Excel export: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Bibliography"

        # Styles
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alt_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")

        # Headers
        headers = ["Year", "Citation Key", "Title", "Authors", "Source", "DOI", "Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Data rows - sorted by year
        sorted_entries = sorted(entries, key=lambda x: x.year or 9999)

        for row_idx, entry in enumerate(sorted_entries, 2):
            cells = [
                (entry.year, "center"),
                (entry.citation_key, None),
                (entry.title, None),
                ("; ".join(entry.authors), None),
                (entry.source, None),
                (entry.doi, None),
                (entry.entry_type, "center"),
            ]

            for col_idx, (value, align) in enumerate(cells, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if align:
                    cell.alignment = Alignment(horizontal=align)

                # Alternating rows
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Column widths
        widths = [8, 18, 55, 35, 40, 25, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Freeze header
        ws.freeze_panes = 'A2'

        wb.save(output_path)

    @staticmethod
    def to_json(entries: List[ExtractedMetadata], output_path: str) -> None:
        """
        Export to JSON format.

        Args:
            entries: List of ExtractedMetadata objects
            output_path: Output file path (.json)
        """
        data = [entry.to_dict() for entry in sorted(entries, key=lambda x: x.year or 9999)]

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)


# Convenience functions for backwards compatibility

def extract_pdf_metadata(
    pdf_path: str,
    gemini_api_key: str,
    existing_metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Convenience function to extract metadata from a PDF.

    Args:
        pdf_path: Path to PDF file
        gemini_api_key: Gemini API key
        existing_metadata: Partial metadata to complete

    Returns:
        Complete metadata dictionary
    """
    extractor = HybridMetadataExtractor(gemini_api_key)
    result = extractor.extract(pdf_path, existing_metadata=existing_metadata)
    return result.to_dict()


def batch_extract_metadata(
    pdf_paths: List[str],
    gemini_api_key: str,
    existing_metadata: Optional[List[Dict[str, Any]]] = None,
    progress_callback: Optional[callable] = None,
    rate_limit_delay: float = 1.0,
) -> List[ExtractedMetadata]:
    """
    Extract metadata from multiple PDFs.

    Args:
        pdf_paths: List of PDF file paths
        gemini_api_key: Gemini API key
        existing_metadata: List of partial metadata dicts (same order as pdf_paths)
        progress_callback: Optional callback(current, total, citation_key)
        rate_limit_delay: Delay between API calls in seconds

    Returns:
        List of ExtractedMetadata objects
    """
    extractor = HybridMetadataExtractor(gemini_api_key)
    existing = existing_metadata or [{}] * len(pdf_paths)

    results = []
    for i, (pdf_path, existing_meta) in enumerate(zip(pdf_paths, existing)):
        if progress_callback:
            key = existing_meta.get("citation_key", Path(pdf_path).stem)
            progress_callback(i + 1, len(pdf_paths), key)

        try:
            result = extractor.extract(pdf_path, existing_metadata=existing_meta)
            results.append(result)
        except Exception as e:
            logger.error(f"Failed to extract metadata from {pdf_path}: {e}")
            # Return a minimal entry on failure
            fallback = ExtractedMetadata(
                citation_key=Path(pdf_path).stem.lower().replace(" ", "_"),
                extraction_method="failed"
            )
            results.append(fallback)

        # Rate limiting for API calls
        if i < len(pdf_paths) - 1:
            time.sleep(rate_limit_delay)

    return results
